#!/usr/bin/python

from pymongo import MongoClient
import sys
import os
import urllib2
import re
import numpy
from scipy import stats
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../clioinfra.js/modules')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../clioinfra.js')))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../')))
from storage import data2store, readdata
from core.configutils import Configuration
import openpyxl
from openpyxl.cell import get_column_letter

def aggregatedata(dbname):
    client = MongoClient()
    db = client.get_database(dbname)
    collection = db.data
    pipe = [ { '$group': { '_id': {'title': "$kbtitle", 'year': "$year", 'spatial': '$kbspatial'}, 'observations': { '$sum': 1 } } }, { '$sort': {'year': 1} } ]
    result = db.data.aggregate(pipeline=pipe)
#    for x in result:
#        print str(x)
    return result

def loadtrackdata(dbname, query):
    client = MongoClient()
    db = client.get_database(dbname)
    collection = db.data
    if query:
	result = db.data.find(query)
    else:
        result = db.data.find()

#    for x in result:
#        print str(x)
    return result

def create_excel_dataset(project, fullpath):
    dbname = "kb%sresult" % project
    trackdbname = "kb%strack" % project
    result = aggregatedata(dbname)

    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.get_active_sheet()
    ws.title = "Dataset"

    i = 1
    order = ["title", "spatial", "year"]
    for item in result:
        j = 1
	if i == 1:
	    ws.column_dimensions["A"].width = 80
	    ws.column_dimensions["B"].width = 20
	    ws.column_dimensions["O"].width = 100
	    ws.column_dimensions["P"].width = 100
	    for value in order:
                c = ws.cell(row=i, column=j)
                c.value = str(value).encode("utf-8")
                j+=1
            c = ws.cell(row=i, column=j)
            c.value = 'observations'
	    j+=1

	# Subjectivity
	data = {}
	query = {"kbtitle": item['_id']['title'], "year": item['_id']['year']}
        trackdata = loadtrackdata(trackdbname, query)
	for record in trackdata:
	    data['Subj average'] = numpy.average(record['subjvector'])
	    data['Subj mode'] = stats.mode(record['subjvector'])[0][0]
	    data['Subj median'] = numpy.median(record['subjvector'])
	    data['Subj min'] = numpy.min(record['subjvector'])
	    data['Subj max'] = numpy.max(record['subjvector'])
	    data['Vector Subjectivity'] = str(record['subjvector'])

            data['Pol average'] = numpy.average(record['polvector'])
            data['Pol mode'] = stats.mode(record['polvector'])[0][0]
            data['Pol median'] = numpy.median(record['polvector'])
            data['Pol min'] = numpy.min(record['polvector'])
            data['Pol max'] = numpy.max(record['polvector'])
	    data['Vector Polarity'] = str(record['polvector'])


	if i == 1:
	    for name in sorted(data):
	        c = ws.cell(row=i, column=j)
                c.value = name
                j+=1	 
	
	# Polarity
	j = 1
	i+=1
        for key in order:
	    value = item['_id'][key]
            c = ws.cell(row=i, column=j)
	    strval = str(value)
	    if strval.isdigit():
		c.value = value
	    else:
                c.value = str(value).encode("utf-8")
            j+=1
	c = ws.cell(row=i, column=j)
	c.value = item['observations']
	j+=1

	# Median, mean, etc
        for name in sorted(data):
            c = ws.cell(row=i, column=j)
            c.value = data[name]
            j+=1

    wb.save(fullpath)
    return fullpath
